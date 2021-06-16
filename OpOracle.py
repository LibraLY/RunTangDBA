# -*- coding: utf-8 -*-
#coding=utf-8
import cx_Oracle
from openpyxl import Workbook

class OpOracle():
    def __init__(self,ora_username,ora_password,ora_host,ora_port,ora_sid):
        '''初始化Oracle连接'''
        self.db=cx_Oracle.connect(ora_username,ora_password,ora_host+':'+ora_port+'/'+ora_sid)
        self.cur=self.db.cursor()
    def Ora_Select(self,strSql):
        '''执行strSql语句进行查询'''
        self.cur.execute(strSql)
        return self.cur.fetchall()
        
    def Ora_Select_List(self,strSql,List):
        '''执行strSql语句进行查询,对应参数使用List中的数据'''
        self.cur.prepare(strSql)
        self.cur.execute(None,List)
        
    def Ora_select_out(self,strSql,List):
        self.cur.prepare(strSql)
        self.cur.execute(None,List)
        fetchall = [x for item in self.cur.fetchall() for x in item]
        fetchall_str = ','.join(str(i) for i in fetchall)
        return fetchall_str
    def Ora_IUD_Single(self,strSql):
        '''执行strSql语句进行增加、删除、修改操作'''
        self.cur.execute(strSql)
        self.db.commit()
    def Ora_IUD_Multi(self,strSql,List):
        '''执行strSql语句进行增加、删除、修改操作，对应参数使用List中的数据'''
        self.cur.prepare(strSql)
        self.cur.executemany(None,List)
        self.db.commit()
    def Ora_excel(self,char):
        wb = Workbook()
        create_wb(wb,char)
        ws = wb.create_sheet(title='',index=0)
        rowcount=1
        list_A_B = []
        for a in self.cur:
            rowcount = rowcount + 1
            list_A_B.append(a)
    
        #读取表字段值
        db_title = [i[0] for i in self.cur.description]
        #遍历表字段值
        for i,description in enumerate(db_title):
            ws.cell(row=1, column = 1 + i).value = description
    
        #读取数据到excel
        for rowNum in range(1,rowcount):
            for i in range(len(self.cur.description)):     
                ws.cell(row=rowNum+1,column=1+i).value=list_A_B[rowNum-1][i]
        wb.save(("C:/Users/Administrator/Desktop/python/数据库大作业/EXCEL/{}.xls".format(char)))

    def Ora_Cur_Close(self):
        '''关闭游标'''
        self.cur.close()
    def Ora_db_Close(self):
        '''关闭Oracle数据库连接'''
        self.db.close()

def create_wb(wb,filename):
    wb.save(("C:/Users/Administrator/Desktop/python/数据库大作业/EXCEL/{}.xls".format(filename)))
    print ("新建Excel："+filename+".xls成功")